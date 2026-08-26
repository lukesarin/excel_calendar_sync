"""
Excel -> Calendar Sync (full mirror: adds AND removes events)
-----------------------------------------------------------------
Makes a calendar match an Excel sheet exactly:
  - Adds events that are in the sheet but not yet in the calendar
  - Removes events that used to be synced but are no longer in the
    sheet (row deleted, or Include changed to N)

SAFETY: only events this tool created are ever touched. Every event it
makes gets an invisible tag in its notes (like "[SyncID:a1b2c3d4]").
On each run, it only deletes tagged events whose ID no longer matches
anything in the spreadsheet - anything else in that calendar (meetings
you were invited to, events you added yourself, etc.) is never touched.

Requires: openpyxl   (install with:  pip3 install openpyxl )

RUN IT:
    python3 excel_to_calendar_gui.py
  or press the Run button in VS Code.
"""

import subprocess
import sys
import os
import re
import uuid
import datetime

try:
    import openpyxl
except ImportError:
    print("This script needs the 'openpyxl' package. Install it with:")
    print("    pip3 install openpyxl")
    sys.exit(1)

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

REQUIRED_COLUMNS = ["Include", "Start Date", "End Date", "Event Title", "Added to Calendar"]
TAG_PATTERN = re.compile(r"\[SyncID:([a-f0-9]{8})\]")


# ---------- AppleScript helpers ----------

def run_applescript(script):
    result = subprocess.run(["osascript", "-e", script], capture_output=True, text=True)
    return result.returncode == 0, result.stdout.strip(), result.stderr.strip()


def get_calendar_names():
    ok, out, err = run_applescript('tell application "Calendar" to name of every calendar')
    if not ok or not out:
        return []
    return [c.strip() for c in out.split(",")]


def get_tagged_events(calendar_name):
    """Returns a dict of {sync_id: title} for every tagged event currently in the calendar."""
    applescript = f'''
    tell application "Calendar"
        tell calendar "{calendar_name}"
            set matchingEvents to every event whose description contains "[SyncID:"
            set output to {{}}
            repeat with e in matchingEvents
                set end of output to ((summary of e) & "@@TITLESEP@@" & (description of e))
            end repeat
        end tell
    end tell
    set AppleScript's text item delimiters to "@@EVENTSEP@@"
    set outputString to output as string
    set AppleScript's text item delimiters to ""
    return outputString
    '''
    ok, out, err = run_applescript(applescript)
    result = {}
    if not ok or not out:
        return result
    for chunk in out.split("@@EVENTSEP@@"):
        if "@@TITLESEP@@" not in chunk:
            continue
        title, desc = chunk.split("@@TITLESEP@@", 1)
        match = TAG_PATTERN.search(desc)
        if match:
            result[match.group(1)] = title.strip()
    return result


def create_event(calendar_name, title, start, end, notes, sync_id):
    def fmt(d):
        return d.strftime("%B %d, %Y")

    tagged_notes = f"{notes}\n[SyncID:{sync_id}]" if notes else f"[SyncID:{sync_id}]"
    escaped_title = title.replace('"', '\\"')
    escaped_notes = tagged_notes.replace('"', '\\"').replace("\n", "\\n")

    applescript = f'''
    tell application "Calendar"
        tell calendar "{calendar_name}"
            set startDate to date "{fmt(start)}"
            set time of startDate to 0
            set endDate to date "{fmt(end)}" + (1 * days)
            set time of endDate to 0
            make new event with properties {{summary:"{escaped_title}", start date:startDate, end date:endDate, allday event:true, description:"{escaped_notes}"}}
        end tell
    end tell
    '''
    ok, out, err = run_applescript(applescript)
    return ok, err


def delete_event_by_sync_id(calendar_name, sync_id):
    applescript = f'''
    tell application "Calendar"
        tell calendar "{calendar_name}"
            delete (every event whose description contains "[SyncID:{sync_id}]")
        end tell
    end tell
    '''
    ok, out, err = run_applescript(applescript)
    return ok, err


# ---------- Excel helpers ----------

def load_sheet(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]

    missing = [c for c in REQUIRED_COLUMNS if c not in headers]
    if missing:
        raise ValueError(f"Missing required column(s): {', '.join(missing)}")

    col = {name: idx for idx, name in enumerate(headers)}
    extra_cols = [h for h in headers if h and h not in REQUIRED_COLUMNS and h != "Notes"]

    to_create = []   # rows with Include=Y, valid dates, no sync ID yet
    keep_ids = set()  # sync IDs that SHOULD still exist (Include=Y rows that already have an ID)

    for row_num in range(2, ws.max_row + 1):
        row = [ws.cell(row=row_num, column=i + 1).value for i in range(len(headers))]
        include = str(row[col["Include"]] or "").strip().upper()
        title = row[col["Event Title"]]
        existing_val = str(row[col["Added to Calendar"]] or "").strip()
        existing_id = existing_val if TAG_PATTERN.fullmatch(f"[SyncID:{existing_val}]") else None

        if include != "Y" or not title:
            continue

        start = row[col["Start Date"]]
        end = row[col["End Date"]]
        if not isinstance(start, datetime.datetime) or not isinstance(end, datetime.datetime):
            continue

        if existing_id:
            keep_ids.add(existing_id)
            continue  # already synced, nothing to create

        notes_parts = []
        if "Notes" in col and row[col["Notes"]]:
            notes_parts.append(str(row[col["Notes"]]))
        for extra in extra_cols:
            val = row[col[extra]]
            if val:
                notes_parts.append(f"{extra}: {val}")

        to_create.append({
            "row_num": row_num,
            "title": str(title).strip(),
            "start": start,
            "end": end,
            "notes": "\n".join(notes_parts),
        })

    return wb, ws, col, to_create, keep_ids


# ---------- GUI ----------

class SyncApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Excel -> Calendar Sync")
        self.geometry("680x560")

        self.excel_path = None
        self.wb = None
        self.ws = None
        self.col = None
        self.to_create = []
        self.to_delete = {}  # sync_id -> title

        pad = {"padx": 12, "pady": 6}

        file_frame = tk.Frame(self)
        file_frame.pack(fill="x", **pad)
        tk.Button(file_frame, text="Choose Excel File...", command=self.pick_file).pack(side="left")
        self.file_label = tk.Label(file_frame, text="No file selected", anchor="w")
        self.file_label.pack(side="left", padx=10)

        cal_frame = tk.Frame(self)
        cal_frame.pack(fill="x", **pad)
        tk.Label(cal_frame, text="Calendar:").pack(side="left")
        self.calendar_var = tk.StringVar()
        self.calendar_dropdown = ttk.Combobox(cal_frame, textvariable=self.calendar_var, state="readonly", width=30)
        self.calendar_dropdown.pack(side="left", padx=10)
        tk.Button(cal_frame, text="Refresh List", command=self.load_calendars).pack(side="left")
        self.calendar_dropdown.bind("<<ComboboxSelected>>", lambda e: self.refresh_preview())

        tk.Label(self, text="Changes that will be made:").pack(anchor="w", padx=12)
        self.tree = ttk.Treeview(self, columns=("action", "dates", "title"), show="headings", height=15)
        self.tree.heading("action", text="Action")
        self.tree.heading("dates", text="Dates")
        self.tree.heading("title", text="Event Title")
        self.tree.column("action", width=70, anchor="center")
        self.tree.column("dates", width=170)
        self.tree.column("title", width=400)
        self.tree.pack(fill="both", expand=True, padx=12, pady=6)

        btn_frame = tk.Frame(self)
        btn_frame.pack(fill="x", **pad)
        self.sync_button = tk.Button(btn_frame, text="Apply Changes", command=self.sync, state="disabled")
        self.sync_button.pack(side="right")

        self.status_label = tk.Label(self, text="", anchor="w", fg="#555555", wraplength=650, justify="left")
        self.status_label.pack(fill="x", padx=12, pady=(0, 10))

        self.load_calendars()

    def load_calendars(self):
        names = get_calendar_names()
        self.calendar_dropdown["values"] = names
        if names:
            self.calendar_dropdown.current(0)
            self.set_status(f"Found {len(names)} calendar(s).")
        else:
            self.set_status("Couldn't find any calendars - open the Calendar app once and try again.")

    def pick_file(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if not path:
            return
        self.excel_path = path
        self.file_label.config(text=os.path.basename(path))
        self.refresh_preview()

    def refresh_preview(self):
        self.tree.delete(*self.tree.get_children())
        if not self.excel_path or not self.calendar_var.get():
            return
        try:
            self.wb, self.ws, self.col, self.to_create, keep_ids = load_sheet(self.excel_path)
        except Exception as e:
            messagebox.showerror("Couldn't read file", str(e))
            self.sync_button.config(state="disabled")
            return

        calendar_name = self.calendar_var.get()
        existing_tagged = get_tagged_events(calendar_name)  # sync_id -> title
        self.to_delete = {sid: title for sid, title in existing_tagged.items() if sid not in keep_ids}

        for e in self.to_create:
            date_str = f"{e['start'].date()} to {e['end'].date()}"
            self.tree.insert("", "end", values=("ADD", date_str, e["title"]))
        for sid, title in self.to_delete.items():
            self.tree.insert("", "end", values=("REMOVE", "", title))

        total_changes = len(self.to_create) + len(self.to_delete)
        if total_changes:
            self.sync_button.config(state="normal")
            self.set_status(f"{len(self.to_create)} event(s) to add, {len(self.to_delete)} to remove.")
        else:
            self.sync_button.config(state="disabled")
            self.set_status("Calendar already matches this spreadsheet - nothing to do.")

    def set_status(self, text):
        self.status_label.config(text=text)

    def sync(self):
        calendar_name = self.calendar_var.get()
        msg = f"Add {len(self.to_create)} event(s) and remove {len(self.to_delete)} event(s) from '{calendar_name}'?"
        if self.to_delete:
            sample = list(self.to_delete.values())[:5]
            msg += "\n\nWill remove:\n" + "\n".join(f"  - {t}" for t in sample)
            if len(self.to_delete) > 5:
                msg += f"\n  ...and {len(self.to_delete) - 5} more"
        if not messagebox.askyesno("Confirm changes", msg):
            return

        added, failed_add = 0, []
        for e in self.to_create:
            sync_id = uuid.uuid4().hex[:8]
            ok, err = create_event(calendar_name, e["title"], e["start"], e["end"], e["notes"], sync_id)
            if ok:
                self.ws.cell(row=e["row_num"], column=self.col["Added to Calendar"] + 1).value = sync_id
                added += 1
            else:
                failed_add.append(e["title"])

        removed, failed_remove = 0, []
        for sid, title in self.to_delete.items():
            ok, err = delete_event_by_sync_id(calendar_name, sid)
            if ok:
                removed += 1
            else:
                failed_remove.append(title)

        self.wb.save(self.excel_path)
        self.set_status(f"Done - added {added}, removed {removed}.")

        problems = []
        if failed_add:
            problems.append("Couldn't add:\n" + "\n".join(failed_add))
        if failed_remove:
            problems.append("Couldn't remove:\n" + "\n".join(failed_remove))
        if problems:
            messagebox.showwarning("Some changes failed", "\n\n".join(problems))
        else:
            messagebox.showinfo("Success", f"Added {added} event(s), removed {removed} event(s).")

        self.refresh_preview()


if __name__ == "__main__":
    try:
        app = SyncApp()
        app.mainloop()
    except tk.TclError as e:
        print("Couldn't open a window - tkinter may not be installed.")
        print(f"\nOriginal error: {e}")
